import os
import requests
from bs4 import BeautifulSoup
import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
import logging
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import messagebox

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Configuration
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE_PATH = '/Users/amran/Documents/Documents/Amran/Fall 2024/Econ 256/Market Journal Fall 2024 Template.xlsx'

CNBC_URLS = {
    "US 10 YR (%)": "https://www.cnbc.com/quotes/US10Y",
    "UK 10 YR (%)": "https://www.cnbc.com/quotes/UK10Y",
    "German 10 YR (%)": "https://www.cnbc.com/quotes/DE10Y",
    "Japan 10 YR (%)": "https://www.cnbc.com/quotes/JP10Y",
    "BRENT CRUDE": "https://www.cnbc.com/quotes/%40LCO.1",
    "DAX": "https://www.cnbc.com/quotes/.GDAXI",
    "BITCOIN": "https://www.cnbc.com/quotes/BTC.CM%3D"
}

# Set your API keys here or use environment variables
FMP_API_KEY = 'yrsO9DP0FmkfbkYksqny5YdxDpgcmRen'  # Replace with your actual FMP API key
NEWS_API_KEY = '4320ca49cf88482e8dbd2b23090629a5'  # Replace with your actual NewsAPI key

def get_fmp_data(symbol):
    base_url = f"https://financialmodelingprep.com/api/v3/quote/{symbol}"
    params = {"apikey": FMP_API_KEY}
    response = requests.get(base_url, params=params)
    data = response.json()
    if data and isinstance(data, list) and len(data) > 0:
        return data[0]['price']
    else:
        logging.error(f"Error fetching data for {symbol}: {data}")
        return None

def get_cnbc_value(url):
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        value_span = soup.find('span', {'class': 'QuoteStrip-lastPrice'})
        if value_span:
            value = value_span.text.strip()
            return float(value.replace(',', '').replace('%', '').strip())
        else:
            logging.error(f"Could not find the value on the page for URL: {url}")
            return None
    except Exception as e:
        logging.error(f"Error fetching data from CNBC for {url}: {str(e)}")
        return None

def get_yf_data():
    yf_tickers = yf.Tickers('^N225 ^FTSE ^DJI ^GSPC GC=F')
    end_date = datetime.now()
    start_date = end_date - timedelta(days=5)
    
    try:
        data = yf_tickers.history(start=start_date, end=end_date)
        if data.empty:
            logging.error("No data returned from yfinance")
            return pd.Series()
        
        latest_data = data['Close'].iloc[-1]
        
        for ticker in latest_data.index:
            if pd.isna(latest_data[ticker]):
                non_nan_values = data['Close'][ticker].dropna()
                if not non_nan_values.empty:
                    latest_data[ticker] = non_nan_values.iloc[-1]
                    logging.warning(f"Used most recent non-NaN value for {ticker}: {latest_data[ticker]}")
                else:
                    logging.error(f"No valid data found for {ticker}")
        
        logging.info(f"Fetched yfinance data: {latest_data}")
        return latest_data
    except Exception as e:
        logging.error(f"Error fetching data from yfinance: {str(e)}")
        return pd.Series()

def get_economic_news():
    base_url = 'https://newsapi.org/v2/everything'
    today = datetime.utcnow().date()
    yesterday = today - timedelta(days=1)
    
    params = {
        'apiKey': NEWS_API_KEY,
        'language': 'en',
        'q': '(economy OR finance OR market) AND (impact OR affect OR influence)',
        'sortBy': 'publishedAt',
        'from': yesterday.isoformat(),
        'to': today.isoformat(),
        'domains': 'wsj.com,reuters.com,bloomberg.com,ft.com,cnbc.com,economist.com',
    }
    
    response = requests.get(base_url, params=params)
    if response.status_code == 200:
        news_data = response.json()
        if news_data.get('articles'):
            for article in news_data['articles']:
                published_at = article.get('publishedAt')
                if published_at:
                    article_date = datetime.strptime(published_at, "%Y-%m-%dT%H:%M:%SZ").date()
                    if article_date >= yesterday:
                        return f"{article['title']} - {article['source']['name']}"
    else:
        logging.error(f"Error fetching news: {response.status_code} - {response.text}")
    
    return "Unable to fetch recent economic news. Please check manually."

def main():
    # Check if API keys are provided
    if not FMP_API_KEY or not NEWS_API_KEY:
        logging.error("API keys are not set. Please set FMP_API_KEY and NEWS_API_KEY.")
        messagebox.showerror("API Key Error", "API keys are not set. Please set FMP_API_KEY and NEWS_API_KEY.")
        return

    # Fetch data
    cnbc_data = {key: get_cnbc_value(url) for key, url in CNBC_URLS.items()}
    closers = get_yf_data()
    fmp_data = {
        'EURO/USD': get_fmp_data("EURUSD"),
        'STG/USD': get_fmp_data("GBPUSD"),
        'USD/YEN': get_fmp_data("USDJPY"),
    }
    economic_news = get_economic_news()
    logging.info(f"Economic News: {economic_news}")

    # Load Excel file
    try:
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active
    except Exception as e:
        logging.error(f"Error loading Excel file: {str(e)}")
        messagebox.showerror("File Error", f"Error loading Excel file: {str(e)}")
        return

    # Find the next empty row
    target_row = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        if row[0].value is None:
            target_row = row[0].row
            break

    if target_row is None:
        target_row = ws.max_row + 1

    # Set today's date in the first column
    today = datetime.now().date()
    ws.cell(row=target_row, column=1, value=today.strftime("%A, %B %d, %Y"))

    logging.info(f"Using row {target_row} for today's date: {today}")

    # Update values
    columns_mapping = {
        'EURO/USD': ('EURO/USD', fmp_data, '{:.4f}'),
        'STG/USD': ('STG/USD', fmp_data, '{:.4f}'),
        'USD/YEN': ('USD/YEN', fmp_data, '{:.2f}'),
        'NIKKEI': ('^N225', closers, '{:.2f}'),
        'DAX ': ('DAX', cnbc_data, '{:.2f}'),
        'FTSE': ('^FTSE', closers, '{:.2f}'),
        'DOW': ('^DJI', closers, '{:.2f}'),
        'S&P': ('^GSPC', closers, '{:.2f}'),
        'US 10 YR (%)': ('US 10 YR (%)', cnbc_data, '{:.2f}%'),
        'GERMAN 10 YR (%)': ('German 10 YR (%)', cnbc_data, '{:.2f}%'),
        'UK 10 YR (%)': ('UK 10 YR (%)', cnbc_data, '{:.2f}%'),
        'JAPAN 10 YR (%)': ('Japan 10 YR (%)', cnbc_data, '{:.2f}%'),
        'GOLD': ('GC=F', closers, '{:.2f}'),
        'BRENT CRUDE ': ('BRENT CRUDE', cnbc_data, '{:.2f}'),
        'BITCOIN ': ('BITCOIN', cnbc_data, '{:.2f}')
    }

    for col in ws.iter_cols(min_col=2, max_col=ws.max_column, min_row=1, max_row=1):
        col_name = col[0].value
        if col_name in columns_mapping:
            key, data_source, format_string = columns_mapping[col_name]
            value = data_source.get(key)
            if value is not None:
                formatted_value = format_string.format(value)
                ws.cell(row=target_row, column=col[0].column, value=formatted_value)
                logging.info(f"Updated {col_name} with value: {formatted_value}")
            else:
                logging.warning(f"No data available for {col_name}")
        elif col_name == "Short DAILY NEWS ITEM(S) that affected one or more of today's prices":
            ws.cell(row=target_row, column=col[0].column, value=economic_news)
            logging.info(f"Updated news column with: {economic_news}")

    # Save the updated Excel file
    try:
        wb.save(EXCEL_FILE_PATH)
        logging.info(f"Excel file updated and saved: {EXCEL_FILE_PATH}")
        messagebox.showinfo("Success", "Excel file updated successfully!")
    except Exception as e:
        logging.error(f"Error saving Excel file: {str(e)}")
        messagebox.showerror("File Error", f"Error saving Excel file: {str(e)}")

def run_app():
    try:
        main()
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Market Data Updater")
    root.geometry("300x150")

    run_button = tk.Button(root, text="Update Market Data", command=run_app, height=2, width=20)
    run_button.pack(pady=40)

    root.mainloop()

"""

import requests
from bs4 import BeautifulSoup
import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
import logging
from datetime import datetime, timedelta

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Configuration
EXCEL_FILE_PATH = '/Users/amran/Documents/Documents/Amran/Fall 2024/Econ 256/Market Journal Fall 2024 Template.xlsx'
CNBC_URLS = {
    "US 10 YR (%)": "https://www.cnbc.com/quotes/US10Y",
    "UK 10 YR (%)": "https://www.cnbc.com/quotes/UK10Y",
    "German 10 YR (%)": "https://www.cnbc.com/quotes/DE10Y",
    "Japan 10 YR (%)": "https://www.cnbc.com/quotes/JP10Y",
    "BRENT CRUDE": "https://www.cnbc.com/quotes/%40LCO.1",
    "DAX": "https://www.cnbc.com/quotes/.GDAXI",
    "BITCOIN": "https://www.cnbc.com/quotes/BTC.CM%3D"
}

FMP_API_KEY = "yrsO9DP0FmkfbkYksqny5YdxDpgcmRen"
NEWS_API_KEY = '4320ca49cf88482e8dbd2b23090629a5'

def get_fmp_data(symbol):
    base_url = f"https://financialmodelingprep.com/api/v3/quote/{symbol}"
    params = {"apikey": FMP_API_KEY}
    response = requests.get(base_url, params=params)
    data = response.json()
    if data and isinstance(data, list) and len(data) > 0:
        return data[0]['price']
    else:
        logging.error(f"Error fetching data for {symbol}: {data}")
        return None

def get_cnbc_value(url):
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        value = soup.find('span', {'class': 'QuoteStrip-lastPrice'}).text.strip()
        return float(value.replace(',', '').replace('%', '').strip())
    except Exception as e:
        logging.error(f"Error fetching data from CNBC for {url}: {str(e)}")
        return None
    
def get_yf_data():
    yf_tickers = yf.Tickers('^N225 ^FTSE ^DJI ^GSPC GC=F')
    end_date = datetime.now()
    start_date = end_date - timedelta(days=5)
    
    try:
        data = yf_tickers.history(start=start_date, end=end_date)
        if data.empty:
            logging.error("No data returned from yfinance")
            return pd.Series()
        
        latest_data = data['Close'].iloc[-1]
        
        for ticker in latest_data.index:
            if pd.isna(latest_data[ticker]):
                non_nan_values = data['Close'][ticker].dropna()
                if not non_nan_values.empty:
                    latest_data[ticker] = non_nan_values.iloc[-1]
                    logging.warning(f"Used most recent non-NaN value for {ticker}: {latest_data[ticker]}")
                else:
                    logging.error(f"No valid data found for {ticker}")
        
        logging.info(f"Fetched yfinance data: {latest_data}")
        return latest_data
    except Exception as e:
        logging.error(f"Error fetching data from yfinance: {str(e)}")
        return pd.Series()

def get_economic_news():
    base_url = 'https://newsapi.org/v2/everything'
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)
    
    params = {
        'apiKey': NEWS_API_KEY,
        'language': 'en',
        'q': '(economy OR finance OR market) AND (impact OR affect OR influence)',
        'sortBy': 'publishedAt',
        'from': yesterday.isoformat(),
        'to': today.isoformat(),
        'domains': 'wsj.com,reuters.com,bloomberg.com,ft.com,cnbc.com,economist.com',
    }
    
    response = requests.get(base_url, params=params)
    if response.status_code == 200:
        news_data = response.json()
        if news_data['articles']:
            for article in news_data['articles']:
                article_date = datetime.strptime(article['publishedAt'], "%Y-%m-%dT%H:%M:%SZ").date()
                if article_date >= yesterday:
                    return f"{article['title']} - {article['source']['name']}"
    
    return "Unable to fetch recent economic news. Please check manually."

# Fetch data
cnbc_data = {key: get_cnbc_value(url) for key, url in CNBC_URLS.items()}
closers = get_yf_data()
fmp_data = {
    'EURO/USD': get_fmp_data("EURUSD"),
    'STG/USD': get_fmp_data("GBPUSD"),
    'USD/YEN': get_fmp_data("USDJPY"),
}
economic_news = get_economic_news()
logging.info(f"Economic News: {economic_news}")

# Load Excel file
wb = load_workbook(EXCEL_FILE_PATH)
ws = wb.active

# Find the next empty row
target_row = None
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    if row[0].value is None:
        target_row = row[0].row
        break

if target_row is None:
    target_row = ws.max_row + 1

# Set today's date in the first column
today = datetime.now().date()
ws.cell(row=target_row, column=1, value=today.strftime("%A, %B %d, %Y"))

logging.info(f"Using row {target_row} for today's date: {today}")

# Update values
columns_mapping = {
    'EURO/USD': ('EURO/USD', fmp_data, '{:.4f}'),
    'STG/USD': ('STG/USD', fmp_data, '{:.4f}'),
    'USD/YEN': ('USD/YEN', fmp_data, '{:.2f}'),
    'NIKKEI': ('^N225', closers, '{:.2f}'),
    'DAX ': ('DAX', cnbc_data, '{:.2f}'),
    'FTSE': ('^FTSE', closers, '{:.2f}'),
    'DOW': ('^DJI', closers, '{:.2f}'),
    'S&P': ('^GSPC', closers, '{:.2f}'),
    'US 10 YR (%)': ('US 10 YR (%)', cnbc_data, '{:.2f}%'),
    'GERMAN 10 YR (%)': ('German 10 YR (%)', cnbc_data, '{:.2f}%'),
    'UK 10 YR (%)': ('UK 10 YR (%)', cnbc_data, '{:.2f}%'),
    'JAPAN 10 YR (%)': ('Japan 10 YR (%)', cnbc_data, '{:.2f}%'),
    'GOLD': ('GC=F', closers, '{:.2f}'),
    'BRENT CRUDE ': ('BRENT CRUDE', cnbc_data, '{:.2f}'),
    'BITCOIN ': ('BITCOIN', cnbc_data, '{:.2f}')
}

for col in ws.iter_cols(min_col=2, max_col=ws.max_column, min_row=1, max_row=1):
    col_name = col[0].value
    if col_name in columns_mapping:
        key, data_source, format_string = columns_mapping[col_name]
        value = data_source.get(key)
        if value is not None:
            formatted_value = format_string.format(value)
            ws.cell(row=target_row, column=col[0].column, value=formatted_value)
            logging.info(f"Updated {col_name} with value: {formatted_value}")
    elif col_name == "Short DAILY NEWS ITEM(S) that affected one or more of today's prices":
        ws.cell(row=target_row, column=col[0].column, value=economic_news)
        logging.info(f"Updated news column with: {economic_news}")

# Save the updated Excel file
wb.save(EXCEL_FILE_PATH)
logging.info(f"Excel file updated and saved: {EXCEL_FILE_PATH}")



import requests
from bs4 import BeautifulSoup
import yfinance as yf
import pandas as pd
from openpyxl import load_workbook
import logging
from datetime import datetime, timedelta

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Configuration
EXCEL_FILE_PATH = '/Users/amran/Documents/Documents/Amran/Fall 2024/Econ 256/Market Journal Fall 2024 Template.xlsx'
CNBC_URLS = {
    "US 10 YR (%)": "https://www.cnbc.com/quotes/US10Y",
    "UK 10 YR (%)": "https://www.cnbc.com/quotes/UK10Y",
    "German 10 YR (%)": "https://www.cnbc.com/quotes/DE10Y",
    "Japan 10 YR (%)": "https://www.cnbc.com/quotes/JP10Y",
    "BRENT CRUDE": "https://www.cnbc.com/quotes/%40LCO.1",
    "DAX": "https://www.cnbc.com/quotes/.GDAXI",
    "BITCOIN": "https://www.cnbc.com/quotes/BTC.CM%3D"
}

FMP_API_KEY = "yrsO9DP0FmkfbkYksqny5YdxDpgcmRen"

def get_fmp_data(symbol):
    base_url = f"https://financialmodelingprep.com/api/v3/quote/{symbol}"
    params = {"apikey": FMP_API_KEY}
    response = requests.get(base_url, params=params)
    data = response.json()
    if data and isinstance(data, list) and len(data) > 0:
        return data[0]['price']
    else:
        logging.error(f"Error fetching data for {symbol}: {data}")
        return None

def get_cnbc_value(url):
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        value = soup.find('span', {'class': 'QuoteStrip-lastPrice'}).text.strip()
        return float(value.replace(',', '').replace('%', '').strip())
    except Exception as e:
        logging.error(f"Error fetching data from CNBC for {url}: {str(e)}")
        return None
    
# Fetch data from yfinance
def get_yf_data():
    yf_tickers = yf.Tickers('^N225 ^FTSE ^DJI ^GSPC GC=F')
    end_date = datetime.now()
    start_date = end_date - timedelta(days=5)  # Fetch data for the last 5 days
    
    try:
        data = yf_tickers.history(start=start_date, end=end_date)
        if data.empty:
            logging.error("No data returned from yfinance")
            return pd.Series()
        
        latest_data = data['Close'].iloc[-1]  # Get the last row of closing prices
        
        # Check for NaN values and try to get the most recent non-NaN value
        for ticker in latest_data.index:
            if pd.isna(latest_data[ticker]):
                non_nan_values = data['Close'][ticker].dropna()
                if not non_nan_values.empty:
                    latest_data[ticker] = non_nan_values.iloc[-1]
                    logging.warning(f"Used most recent non-NaN value for {ticker}: {latest_data[ticker]}")
                else:
                    logging.error(f"No valid data found for {ticker}")
        
        logging.info(f"Fetched yfinance data: {latest_data}")
        return latest_data
    except Exception as e:
        logging.error(f"Error fetching data from yfinance: {str(e)}")
        return pd.Series()

# Fetch data
cnbc_data = {key: get_cnbc_value(url) for key, url in CNBC_URLS.items()}
closers = get_yf_data()
fmp_data = {
    'EURO/USD': get_fmp_data("EURUSD"),
    'STG/USD': get_fmp_data("GBPUSD"),
    'USD/YEN': get_fmp_data("USDJPY"),
}

# Load Excel file
wb = load_workbook(EXCEL_FILE_PATH)
ws = wb.active

# Find the row for today's date
today = datetime.now().date()
target_row = None
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    cell_value = row[0].value
    if isinstance(cell_value, datetime):
        if cell_value.date() == today:
            target_row = row[0].row
            break
    elif isinstance(cell_value, str):
        try:
            date = datetime.strptime(cell_value, "%A, %B %d, %Y").date()
            if date == today:
                target_row = row[0].row
                break
        except ValueError:
            pass

if target_row is None:
    logging.error("No row found for today's date.")
    exit()

# Update values
columns_mapping = {
    'EURO/USD': ('EURO/USD', fmp_data, '{:.4f}'),
    'STG/USD': ('STG/USD', fmp_data, '{:.4f}'),
    'USD/YEN': ('USD/YEN', fmp_data, '{:.2f}'),
    'NIKKEI': ('^N225', closers, '{:.2f}'),
    'DAX ': ('DAX', cnbc_data, '{:.2f}'),
    'FTSE': ('^FTSE', closers, '{:.2f}'),
    'DOW': ('^DJI', closers, '{:.2f}'),
    'S&P': ('^GSPC', closers, '{:.2f}'),
    'US 10 YR (%)': ('US 10 YR (%)', cnbc_data, '{:.2f}%'),
    'GERMAN 10 YR (%)': ('German 10 YR (%)', cnbc_data, '{:.2f}%'),
    'UK 10 YR (%)': ('UK 10 YR (%)', cnbc_data, '{:.2f}%'),
    'JAPAN 10 YR (%)': ('Japan 10 YR (%)', cnbc_data, '{:.2f}%'),
    'GOLD': ('GC=F', closers, '{:.2f}'),
    'BRENT CRUDE ': ('BRENT CRUDE', cnbc_data, '{:.2f}'),
    'BITCOIN ': ('BITCOIN', cnbc_data, '{:.2f}')
}

for col in ws.iter_cols(min_col=2, max_col=ws.max_column, min_row=1, max_row=1):
    col_name = col[0].value
    if col_name in columns_mapping:
        key, data_source, format_string = columns_mapping[col_name]
        value = data_source.get(key)
        if value is not None:
            formatted_value = format_string.format(value)
            ws.cell(row=target_row, column=col[0].column, value=formatted_value)
            logging.info(f"Updated {col_name} with value: {formatted_value}")

# Save the updated Excel file
wb.save(EXCEL_FILE_PATH)
logging.info(f"Excel file updated and saved: {EXCEL_FILE_PATH}")
"""