import os
import uuid
import json
import requests
import time
from bs4 import BeautifulSoup
import yfinance as yf
import pandas as pd
from openpyxl import load_workbook, Workbook
import logging
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import messagebox, simpledialog
from cryptography.fernet import Fernet
import sys

# ---------------------- Configuration ---------------------- #

# Get the user's home directory
home_dir = os.path.expanduser('~')

# Determine the Downloads directory
downloads_dir = os.path.join(home_dir, 'Downloads')

# Ensure the Downloads directory exists
if not os.path.exists(downloads_dir):
    logging.error(f"Downloads directory not found at {downloads_dir}.")
    # Attempt to create the Downloads directory
    try:
        os.makedirs(downloads_dir)
        logging.info(f"Downloads directory created at {downloads_dir}.")
    except Exception as e:
        logging.error(f"Failed to create Downloads directory: {str(e)}")
        sys.exit(1)

# File paths
EXCEL_FILE_PATH = os.path.join(downloads_dir, 'Market Journal.xlsx')
LICENSE_KEY_FILE = os.path.join(home_dir, '.license_key')  # Encrypted license key
API_KEY_FILE = os.path.join(home_dir, '.api_key')
ENCRYPTION_KEY_FILE = os.path.join(home_dir, '.encryption_key')

# License server URL
LICENSE_SERVER_URL = "http://localhost:5001"

# CNBC URLs for data fetching
CNBC_URLS = {
    "US 10 YR (%)": "https://www.cnbc.com/quotes/US10Y",
    "UK 10 YR (%)": "https://www.cnbc.com/quotes/UK10Y",
    "German 10 YR (%)": "https://www.cnbc.com/quotes/DE10Y",
    "Japan 10 YR (%)": "https://www.cnbc.com/quotes/JP10Y",
    "BRENT CRUDE": "https://www.cnbc.com/quotes/%40LCO.1",
    "DAX": "https://www.cnbc.com/quotes/.GDAXI",
    "BITCOIN": "https://www.cnbc.com/quotes/BTC.CM%3D",
}

FMP_API_KEY = ''  # Will be set by loading from file or user input

# ---------------------- Logging Setup ---------------------- #

logging.basicConfig(
    level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s'
)

# ---------------------- Helper Functions ---------------------- #

def get_device_id():
    """
    Retrieve the device's unique identifier (UUID).
    """
    # Use uuid.getnode() to get the MAC address
    mac = uuid.getnode()
    return str(mac)

def generate_and_save_encryption_key():
    """
    Generate a new encryption key and save it to a file.
    """
    key = Fernet.generate_key()
    try:
        with open(ENCRYPTION_KEY_FILE, 'wb') as f:
            f.write(key)
        os.chmod(ENCRYPTION_KEY_FILE, 0o600)  # Read-write for owner only
        logging.info("Encryption key generated and saved.")
        return key
    except Exception as e:
        logging.error(f"Error generating/saving encryption key: {str(e)}")
        return None

def load_encryption_key():
    """
    Load the encryption key from the file.
    """
    if os.path.exists(ENCRYPTION_KEY_FILE):
        try:
            with open(ENCRYPTION_KEY_FILE, 'rb') as f:
                key = f.read()
            logging.info("Encryption key loaded.")
            return key
        except Exception as e:
            logging.error(f"Error loading encryption key: {str(e)}")
            return None
    else:
        # Encryption key does not exist, generate and save one
        key = generate_and_save_encryption_key()
        return key

def validate_license_with_server(license_key, device_id):
    """
    Validate the license key with the Flask server.
    """
    url = f"{LICENSE_SERVER_URL}/api/validate_license"
    data = {
        "license_key": license_key,
        "device_id": device_id
    }
    try:
        response = requests.post(url, json=data, timeout=10)
        response.raise_for_status()
        return response.json()
    except requests.RequestException as e:
        if isinstance(e, requests.ConnectionError):
            logging.error(f"Unable to connect to license server: {str(e)}")
        elif isinstance(e, requests.Timeout):
            logging.error(f"License server request timed out: {str(e)}")
        elif isinstance(e, requests.HTTPError):
            logging.error(f"HTTP error occurred: {e.response.status_code} - {e.response.text}")
        else:
            logging.error(f"An unexpected error occurred: {str(e)}")
        return None

def is_license_server_accessible():
    """
    Check if the license server is accessible.
    """
    test_url = f"{LICENSE_SERVER_URL}/test"
    try:
        response = requests.get(test_url, timeout=5)
        response.raise_for_status()
        return True
    except requests.RequestException as e:
        logging.error(f"License server accessibility check failed: {str(e)}")
        return False

def is_device_authorized(max_attempts=3):
    """
    Check if the current device is authorized to run the application.
    Allows for multiple attempts to enter a valid license key.
    """
    current_device_id = get_device_id()

    for attempt in range(max_attempts):
        # Check if a license key is already saved locally
        if os.path.exists(LICENSE_KEY_FILE):
            try:
                with open(LICENSE_KEY_FILE, 'rb') as f:
                    encrypted_license = f.read()
                decrypted_license = decrypt_data(encrypted_license)
                # Optionally, you can re-validate the license with the server
                validation_result = validate_license_with_server(decrypted_license, current_device_id)
                if validation_result and validation_result.get('valid'):
                    logging.info("Device authorized with existing license key.")
                    return True
                else:
                    logging.warning("Existing license key is invalid or expired.")
            except Exception as e:
                logging.error(f"Error reading license key file: {str(e)}")

        # Prompt for license key
        license_key = simpledialog.askstring("License Key", f"Please enter your license key (Attempt {attempt+1}/{max_attempts}):", parent=root)
        if not license_key:
            logging.error("No license key provided.")
            if attempt < max_attempts - 1:
                continue
            else:
                messagebox.showerror("License Key Error", "No license key provided. Exiting application.")
                return False

        # Validate the license key with the server
        validation_result = validate_license_with_server(license_key, current_device_id)
        if validation_result is None:
            messagebox.showerror("Server Error", "Unable to communicate with the license server. Please try again later.")
            return False

        if validation_result.get('valid'):
            # Save the valid license key locally
            try:
                save_license_key(license_key)
                logging.info("Device authorized and license key registered.")
                messagebox.showinfo("Success", "License key validated and registered successfully.")
                return True
            except Exception as e:
                logging.error(f"Error saving license key: {str(e)}")
                messagebox.showerror("Error", "Failed to save license key. Exiting application.")
                return False
        else:
            logging.error(f"License validation failed: {validation_result.get('message')}")
            if attempt < max_attempts - 1:
                messagebox.showerror("License Key Error", f"License validation failed: {validation_result.get('message')}. Please try again.")
            else:
                messagebox.showerror("License Key Error", f"License validation failed: {validation_result.get('message')}. Exiting application.")

    return False

def save_license_key(license_key):
    """
    Encrypt and save the license key to a file.
    """
    encrypted_key = encrypt_data(license_key)
    with open(LICENSE_KEY_FILE, 'wb') as f:
        f.write(encrypted_key)
    os.chmod(LICENSE_KEY_FILE, 0o600)  # Read-write for owner only
    logging.info("License key saved locally.")

def encrypt_data(data):
    """
    Encrypt data using Fernet.
    """
    encrypted_data = fernet.encrypt(data.encode('utf-8'))
    return encrypted_data

def decrypt_data(encrypted_data):
    """
    Decrypt data using Fernet.
    """
    decrypted_data = fernet.decrypt(encrypted_data).decode('utf-8')
    return decrypted_data

def save_api_key(api_key):
    """
    Encrypt and save the FMP API key to a file.
    """
    encrypted_key = encrypt_data(api_key)
    try:
        with open(API_KEY_FILE, 'wb') as f:
            f.write(encrypted_key)
        os.chmod(API_KEY_FILE, 0o600)  # Read-write for owner only
        logging.info("API key saved.")
    except Exception as e:
        logging.error(f"Error saving API key: {str(e)}")

def load_api_key():
    """
    Load and decrypt the FMP API key from the file.
    """
    if os.path.exists(API_KEY_FILE):
        try:
            with open(API_KEY_FILE, 'rb') as f:
                encrypted_key = f.read()
            api_key = decrypt_data(encrypted_key)
            logging.info("API key loaded.")
            return api_key
        except Exception as e:
            logging.error(f"Error loading API key: {str(e)}")
            return None
    else:
        return None

def create_excel_file():
    """
    Create the Excel file with predefined headers.
    """
    try:
        wb = Workbook()
        ws = wb.active

        # Define your headers
        headers = ['Date', 'EURO/USD', 'STG/USD', 'USD/YEN', 'NIKKEI', 'DAX ', 'FTSE',
                   'DOW', 'S&P', 'US 10 YR (%)', 'GERMAN 10 YR (%)', 'UK 10 YR (%)',
                   'JAPAN 10 YR (%)', 'GOLD', 'BRENT CRUDE ', 'BITCOIN ']

        ws.append(headers)
        wb.save(EXCEL_FILE_PATH)
        logging.info(f"Excel file created at {EXCEL_FILE_PATH}")
    except Exception as e:
        logging.error(f"Error creating Excel file: {str(e)}")

def get_fmp_data(symbol):
    """
    Fetch data from Financial Modeling Prep API.
    """
    base_url = f"https://financialmodelingprep.com/api/v3/quote/{symbol}"
    params = {"apikey": FMP_API_KEY}
    try:
        response = requests.get(base_url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        if data and isinstance(data, list) and len(data) > 0:
            return data[0]['price']
        else:
            logging.error(f"Error fetching data for {symbol}: {data}")
            return None
    except Exception as e:
        logging.error(f"Error fetching data for {symbol}: {str(e)}")
        return None

def get_cnbc_value(url):
    """
    Scrape data from CNBC.
    """
    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        value_span = soup.find('span', {'class': 'QuoteStrip-lastPrice'})
        if value_span:
            value = value_span.text.strip()
            # Remove any percent signs and commas
            value_clean = value.replace(',', '').replace('%', '').strip()
            # Attempt to convert to float
            try:
                return float(value_clean)
            except ValueError:
                logging.error(f"Value conversion error for {url}: {value_clean}")
                return None
        else:
            logging.error(f"Could not find the value on the page for URL: {url}")
            return None
    except Exception as e:
        logging.error(f"Error fetching data from CNBC for {url}: {str(e)}")
        return None

def get_yf_data():
    """
    Fetch data using yfinance.
    """
    yf_tickers = yf.Tickers('^N225 ^FTSE ^DJI ^GSPC GC=F')
    end_date = datetime.now()
    start_date = end_date - timedelta(days=5)

    try:
        data = yf_tickers.history(start=start_date, end=end_date)
        if data.empty:
            logging.error("No data returned from yfinance")
            return pd.Series()

        latest_data = data['Close'].iloc[-1]

        for ticker in latest_data.index:
            if pd.isna(latest_data[ticker]):
                non_nan_values = data['Close'][ticker].dropna()
                if not non_nan_values.empty:
                    latest_data[ticker] = non_nan_values.iloc[-1]
                    logging.warning(f"Used most recent non-NaN value for {ticker}: {latest_data[ticker]}")
                else:
                    logging.error(f"No valid data found for {ticker}")

        logging.info(f"Fetched yfinance data: {latest_data}")
        return latest_data
    except Exception as e:
        logging.error(f"Error fetching data from yfinance: {str(e)}")
        return pd.Series()

# ---------------------- Main Function ---------------------- #

def main():
    # Check if license server is accessible
    if not is_license_server_accessible():
        messagebox.showerror("Server Error", "License server is not accessible. Please check your connection and try again.")
        root.destroy()
        sys.exit(1)

    # Check device authorization
    if not is_device_authorized():
        root.destroy()  # Close the application
        sys.exit(1)

    global FMP_API_KEY

    # Load the API key if it exists
    FMP_API_KEY = load_api_key()

    # Prompt user for FMP API key if not set
    if not FMP_API_KEY:
        FMP_API_KEY = simpledialog.askstring("Input", "Please enter your FMP API key:", parent=root)
        if not FMP_API_KEY:
            logging.error("No API key provided.")
            messagebox.showerror("API Key Error", "No API key provided.")
            root.destroy()  # Close the application
            sys.exit(1)
        else:
            save_api_key(FMP_API_KEY)

    # Check if Excel file exists, if not create it
    if not os.path.exists(EXCEL_FILE_PATH):
        create_excel_file()
        logging.info(f"Created new Excel file: {EXCEL_FILE_PATH}")

    # Fetch data
    cnbc_data = {key: get_cnbc_value(url) for key, url in CNBC_URLS.items()}
    closers = get_yf_data()
    fmp_data = {
        'EURO/USD': get_fmp_data("EURUSD"),
        'STG/USD': get_fmp_data("GBPUSD"),
        'USD/YEN': get_fmp_data("USDJPY"),
    }

    # Load Excel file
    try:
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active
    except Exception as e:
        logging.error(f"Error loading Excel file: {str(e)}")
        messagebox.showerror("File Error", f"Error loading Excel file: {str(e)}")
        root.destroy()  # Close the application
        sys.exit(1)

    # Find the next empty row
    target_row = ws.max_row + 1

    # Set today's date in the first column
    today = datetime.now().date()
    ws.cell(row=target_row, column=1, value=today.strftime("%A, %B %d, %Y"))

    logging.info(f"Using row {target_row} for today's date: {today}")

    # Update values
    columns_mapping = {
        'EURO/USD': ('EURO/USD', fmp_data, '{:.4f}'),
        'STG/USD': ('STG/USD', fmp_data, '{:.4f}'),
        'USD/YEN': ('USD/YEN', fmp_data, '{:.2f}'),
        'NIKKEI': ('^N225', closers, '{:.2f}'),
        'DAX ': ('DAX', cnbc_data, '{:.2f}'),
        'FTSE': ('^FTSE', closers, '{:.2f}'),
        'DOW': ('^DJI', closers, '{:.2f}'),
        'S&P': ('^GSPC', closers, '{:.2f}'),
        'US 10 YR (%)': ('US 10 YR (%)', cnbc_data, '{:.2f}%'),
        'GERMAN 10 YR (%)': ('German 10 YR (%)', cnbc_data, '{:.2f}%'),
        'UK 10 YR (%)': ('UK 10 YR (%)', cnbc_data, '{:.2f}%'),
        'JAPAN 10 YR (%)': ('Japan 10 YR (%)', cnbc_data, '{:.2f}%'),
        'GOLD': ('GC=F', closers, '{:.2f}'),
        'BRENT CRUDE ': ('BRENT CRUDE', cnbc_data, '{:.2f}'),
        'BITCOIN ': ('BITCOIN', cnbc_data, '{:.2f}'),
    }

    for col in ws.iter_cols(min_col=2, max_col=ws.max_column, min_row=1, max_row=1):
        col_name = col[0].value
        if col_name in columns_mapping:
            key, data_source, format_string = columns_mapping[col_name]
            value = data_source.get(key)
            if value is not None:
                formatted_value = format_string.format(value)
                ws.cell(row=target_row, column=col[0].column, value=formatted_value)
                logging.info(f"Updated {col_name} with value: {formatted_value}")
            else:
                logging.warning(f"No data available for {col_name}")

    # Save the updated Excel file
    try:
        wb.save(EXCEL_FILE_PATH)
        logging.info(f"Excel file updated and saved: {EXCEL_FILE_PATH}")
        messagebox.showinfo("Success", f"Excel file updated successfully!\nLocation: {EXCEL_FILE_PATH}")
        root.destroy()  # Close the application after success
        sys.exit(0)
    except Exception as e:
        logging.error(f"Error saving Excel file: {str(e)}")
        messagebox.showerror("File Error", f"Error saving Excel file: {str(e)}")
        root.destroy()  # Close the application on error
        sys.exit(1)

def run_app():
    try:
        main()
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")
        messagebox.showerror("Error", f"An error occurred: {str(e)}")
        root.destroy()  # Close the application on unexpected error
        sys.exit(1)

# ---------------------- GUI Setup ---------------------- #

if __name__ == "__main__":
    # Load or generate the encryption key
    encryption_key = load_encryption_key()
    if encryption_key is None:
        messagebox.showerror("Encryption Key Error", "Failed to load or generate encryption key.")
        sys.exit(1)

    fernet = Fernet(encryption_key)

    root = tk.Tk()
    root.title("Market Data Updater")
    root.geometry("300x150")

    run_button = tk.Button(
        root, text="Update Market Data", command=run_app, height=2, width=20
    )
    run_button.pack(pady=40)

    root.mainloop()