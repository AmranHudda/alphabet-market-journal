import requests
from openpyxl import load_workbook
from datetime import datetime, timedelta

EXCEL_FILE_PATH = '/Users/amran/Documents/Documents/Amran/Fall 2024/Econ 256/Market Journal Fall 2024 Template.xlsx'
NEWS_API_KEY = '4320ca49cf88482e8dbd2b23090629a5'

def get_economic_news():
    base_url = 'https://newsapi.org/v2/everything'
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)
    
    params = {
        'apiKey': NEWS_API_KEY,
        'language': 'en',
        'q': 'economy OR finance OR market',
        'sortBy': 'relevancy',
        'from': yesterday.isoformat(),
        'to': today.isoformat(),
        'domains': 'wsj.com,reuters.com,bloomberg.com,ft.com,cnbc.com,economist.com',
    }
    
    response = requests.get(base_url, params=params)
    if response.status_code == 200:
        news_data = response.json()
        if news_data['articles']:
            article = news_data['articles'][0]  # Get the top article
            return f"{article['title']} - {article['source']['name']}"
    
    # If no news found, broaden the search
    params['q'] = 'business OR stock OR trade'
    response = requests.get(base_url, params=params)
    if response.status_code == 200:
        news_data = response.json()
        if news_data['articles']:
            article = news_data['articles'][0]  # Get the top article
            return f"{article['title']} - {article['source']['name']}"
    
    return "Unable to fetch economic news. Please check manually."

# Fetch economic news
economic_news = get_economic_news()
print("Economic News:")
print(economic_news)

# Load the existing Excel file
wb = load_workbook(EXCEL_FILE_PATH)
ws = wb.active

# Find the row for today's date
today = datetime.now().date()
target_row = None
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    cell_date = row[0].value
    if isinstance(cell_date, datetime):
        if cell_date.date() == today:
            target_row = row[0].row
            break

if target_row is None:
    print("No row found for today's date.")
else:
    # Update the economic news in the Excel sheet
    for col in ws.iter_cols(min_col=2, max_col=ws.max_column, min_row=1, max_row=1):
        col_name = col[0].value
        if col_name == "Short DAILY NEWS ITEM(S) that affected one or more of today's prices":
            ws.cell(row=target_row, column=col[0].column, value=economic_news)
            print(f"Updated news column with: {economic_news}")
            break
    else:
        print("News column not found in the Excel sheet.")

    # Save the updated Excel file
    wb.save(EXCEL_FILE_PATH)
    print(f"Excel file updated and saved: {EXCEL_FILE_PATH}")